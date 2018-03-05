from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from style import style_range
from funcModules import getTargetCell
from openpyxl.utils import get_column_letter

# 參數設定區
strFileName = '11月工作表'  # input('請輸入檔案名稱：')
strDistinctCol = '牙醫診所'  # input('請輸入要建立分頁所依據的欄位名稱：')
strHeadText = '2017 年 11月 請 款 單'  # input('請輸入表頭文字：')
strOutPutFileName = strFileName + '_output'
strTotalPriceColName = '總價'
# strTail = 'PS:\r\n一.資料若有錯誤，請立即通知歐美技工所\r\n二.未寫技工單者,以收單日期填寫\r\n三.9月份後請款單，以當月收單日期為主\r\n四.107年1月部分產品調漲通知\r\n五.新增請款單方式:郵寄，mail，line\r\n\t\tmail:k62718@gmail.com\r\n\t\tline ID:0953162482'
strTail = '''PS:
一.資料若有錯誤，請立即通知歐美技工所
二.未寫技工單者,以收單日期填寫
三.9月份後請款單，以當月收單日期為主
四.107年1月部分產品調漲通知
五.新增請款單方式:郵寄，mail，line
\t\tmail:k62718@gmail.com
\t\tline ID:0953162482'''

# 樣式
fontHead = Font(b=True, color="000000", size=28)  # 粗體, 28號字, 黑色
fontTotalPrice = Font(b=True, color='000000', size=14)  # 粗體, 14號字, 黑色
fontTail = Font(b=False, color='000000', size=14)  #
alHVCenter = Alignment(horizontal="center", vertical="center")  # 水平垂直置中排列
alHCenter = Alignment(horizontal="center", vertical="bottom")  # 水平置中排列
alHleftVBottom = Alignment(horizontal="left", vertical="bottom")  # 水平置中排列
alHleftVTop = Alignment(horizontal="left", vertical="top", wrap_text=True)  # 水平置中排列文字換行
noborder = Side(border_style="thin", color="FFFFFF")  # 沒有框
redThickBorder = Side(border_style='thick', color='FF0000')  # 紅粗框
blackThickBorder = Side(border_style='thick', color='000000')  # 黑粗框
grayThinBorder = Side(border_style='thin', color='969696')  # 灰細框
borderHead = Border(top=noborder, left=noborder, right=noborder, bottom=noborder)
borderColLists = Border(top=blackThickBorder, left=noborder, right=noborder, bottom=blackThickBorder)
borderTotalPrice = Border(top=blackThickBorder, left=blackThickBorder, right=blackThickBorder, bottom=blackThickBorder)
borderReportCell = Border(top=grayThinBorder, left=noborder, right=noborder, bottom=grayThinBorder)
borderTail = Border(top=grayThinBorder, left=grayThinBorder, right=grayThinBorder, bottom=grayThinBorder)

# 讀入檔案
wb = load_workbook(filename=f'{strFileName}.xlsx')
sheet_ranges = wb['總表']

# 找到要distinct的欄位是在第幾欄
listCols = list(sheet_ranges['1']) # 載入欄位列
listCols = list(filter(lambda x: x.value is not None, listCols))  # 去掉值為None的cell
idxLastCol = listCols[len(listCols)-1].column  # 找到最後一個欄位的英文index
FirstCellInColDistinct = getTargetCell(strDistinctCol, listCols)  # 找到要distinct的欄位是在那一格
# 找到放總金額數字的欄的第一格
FirstCellInColTotalPrice = getTargetCell(strTotalPriceColName, listCols)
# 總金額文字合併格的第一欄和最後一欄
firstColInTotalPriceText = get_column_letter(FirstCellInColTotalPrice.col_idx - 3)
lastColInTotalPriceText = get_column_letter(FirstCellInColTotalPrice.col_idx - 1)

listColNames = list(map(lambda x: x.value, listCols))  # 抓出所有欄名形成list

# 讀入要distinct的col
listDistinctCol = sheet_ranges[FirstCellInColDistinct.column]
listDistinctCol = list(filter(lambda x: x.value is not None, listDistinctCol))  # 去掉值為None的cell
listDistinctCol = list(map(lambda x: x.value, listDistinctCol))  # 把每個cell的值都取出來形成list
numRows = len(listDistinctCol)  # 設定有幾列
listDistinctCol.remove(strDistinctCol)  # 移除欄位cell
listDistinctCol = list(set(listDistinctCol))  # distinct欄位值

# 在記憶體中建立新的excel檔案，並且依照listDistictCol不同的cell值建立sheets
wbOutPut = Workbook()
listSheets = list(map(lambda x: wbOutPut.create_sheet(x, 0), listDistinctCol))  # 在記憶體中要輸出的新excel檔裡建立sheets

# 寫入表頭
for sheet in listSheets:
    cellHead = sheet['A1']
    cellHead.value = strHeadText
    style_range(sheet, f'A1:{idxLastCol}3', border=borderHead, fill='', font=fontHead, alignment=alHVCenter)
    sheet.append(listColNames)

# 讀取每一列，根據strDistinctCol的值把列塞到不同的sheet
dictSheetIndex = {}  # 建立sheet字典
idx = 0
for item in listDistinctCol:  # 給字典塞入值，這個數字要用在sheet的索引
    dictSheetIndex[item] = idx
    idx += 1

# 判斷是哪間診所，分配到不同sheet
for numRowIdx in range(2, numRows + 1):
    row = list(sheet_ranges[f'A{numRowIdx}:{idxLastCol}{numRowIdx}'][0])  # 把列讀進來
    idxSheet = dictSheetIndex[sheet_ranges[f'{FirstCellInColDistinct.column}{numRowIdx}'].value]  # 根據要distinct的目標欄位的值，找出要寫入的sheet index
    row = list(map(lambda x: x.value, row))  # 把列的所有cell的值取出來重組成list
    listSheets[idxSheet].append(row)

# 計算總價格還有格線設定、表尾
for sheet in listSheets:
    numRowCount = sheet.max_row
    numColCount = sheet.max_column
    sheet.cell(row = numRowCount + 1, column = sheet[f'{FirstCellInColTotalPrice.column}{numRowCount + 1}'].col_idx - 3, value = '總金額')  # 在某個格子中加入"總金額"三個字
    style_range(sheet, f'{firstColInTotalPriceText}{numRowCount + 1}:{lastColInTotalPriceText}{numRowCount + 1}', border=borderTotalPrice, fill='', font=fontTotalPrice, alignment=alHVCenter)
    # 計算總價格
    sheet[f'{FirstCellInColTotalPrice.column}{numRowCount + 1}'] = f'=SUM({FirstCellInColTotalPrice.column}4:{FirstCellInColTotalPrice.column}{numRowCount})'  # 把總價加總一下
    # 畫格線，要一格格畫
    for numRowIdx in range(4, numRowCount + 1):
        for numColIdx in range(1,numColCount + 1):
            if numRowIdx == 4:
                sheet[f'{get_column_letter(numColIdx)}{numRowIdx}'].border = borderColLists
                sheet[f'{get_column_letter(numColIdx)}{numRowIdx}'].alignment = alHCenter
            else:
                sheet[f'{get_column_letter(numColIdx)}{numRowIdx}'].border = borderReportCell
                sheet[f'{get_column_letter(numColIdx)}{numRowIdx}'].alignment = alHCenter
    # 加表尾
    sheet[f'A{numRowCount + 3}'] = strTail
    # 設定表尾的格式
    style_range(sheet, f'A{numRowCount + 3}:{get_column_letter(numColCount)}{numRowCount + 3 + 10}', border=borderTail, fill='', font=fontTail, alignment=alHleftVTop)

# 寫入excel檔案
wbOutPut.save(filename = f'{strOutPutFileName}.xlsx')

